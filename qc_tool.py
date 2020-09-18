from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, Alignment
import os

def get_glossary(glossary):
    output = set()
    row_num = 1
    while True:
        source_text = glossary[f"A{row_num}"].value
        target_text = glossary[f"B{row_num}"].value
        
        if source_text is None and target_text is None:
            break

        output.add((source_text, target_text))
        row_num += 1
    
    return output
    
def fix_fuzzy(content):
    fuzzy_font = Font(name="Calibri", size=11, bold=True, color=Color(rgb="FFFF9900"))
    exact_font = Font(name="Calibri", size=11, bold=True, color=Color(rgb="FF7AA221"))
    rep_font = Font(name="Calibri", size=11, bold=True, color=Color(rgb="FFF9C724"))
    
    if "R" in content:
        return rep_font
    
    elif "100" in content:
        return exact_font
    
    elif "%" in content:
        return fuzzy_font
    
    return None

def qc_tool(source_file, glossary_file, output_file="output.xlsx"):
    # Todo: Sheet names standardization?
    # This is a formatting used when generating the original spreadsheet, we name the glossary file CaseXXX with a capital C, so we need to replace it accordingly.
    name = f.split("_")[1].replace('c', 'C')
    g_name = "glossary"

    wb = load_workbook(filename=source_file)
    qc_case = wb[name]

    glossary_wb = load_workbook(filename=glossary_file)
    glossary = glossary_wb[g_name]

    row_num = 4
    bugged = set()
    comments = {}

    while True:
        segment_id = qc_case[f"A{row_num}"].value
        source_text = qc_case[f"B{row_num}"].value
        target_text = qc_case[f"C{row_num}"].value

        if segment_id is None and source_text is None and target_text is None:
            break

        glossary_ = get_glossary(glossary)

        for source, target in glossary_:
            if source in source_text:
                if not target in target_text:
                    bugged.add(f"C{row_num}")
                    try:
                        comments[f"G{row_num}"] += f"[Source: {source} | Target: {target}]\n"
                    except KeyError:
                        comments[f"G{row_num}"] = f"[Source: {source} | Target: {target}]\n"

        new_fuzzy_font = fix_fuzzy(qc_case[f"D{row_num}"].value)

        if new_fuzzy_font:
            qc_case[f"D{row_num}"].font = new_fuzzy_font

        row_num += 1

    for i in bugged:
        qc_case[i].fill = filling

    qc_case.column_dimensions['G'].width = 40

    for k,v in comments.items():
        qc_case[k] = v
        qc_case[k].alignment = Alignment(wrapText=True)

    wb.save(output_file)
